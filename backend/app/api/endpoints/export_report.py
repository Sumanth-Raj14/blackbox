"""Export/Reporting endpoints - real XLSX and PDF generation with tenant isolation."""

import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.db.session import get_db
from app.models.user import User

router = APIRouter()


def _build_header(ws, headers):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


@router.get("/parts/xlsx")
async def export_parts_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook

    result = await db.execute(
        text(
            'SELECT pn, name, description, category, "subCategory", manufacturer, vendor, cost, lead, origin, status, material, weight, mpn, "htsCode", "unspscCode" FROM parts WHERE "tenantId" = :tid ORDER BY pn'
        ),
        {"tid": current_user.tenantId},
    )
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"

    _build_header(
        ws,
        [
            "PN",
            "Name",
            "Description",
            "Category",
            "Sub-Category",
            "Manufacturer",
            "Vendor",
            "Cost (USD)",
            "Lead (days)",
            "Origin",
            "Status",
            "Material",
            "Weight (g)",
            "MPN",
            "HTS Code",
            "UNSPSC Code",
        ],
    )

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parts_export.xlsx"},
    )


@router.get("/vendors/xlsx")
async def export_vendors_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook

    result = await db.execute(
        text(
            'SELECT name, contact, email, phone, address, category, rating, status, "paymentTerms" FROM vendors WHERE "tenantId" = :tid ORDER BY name'
        ),
        {"tid": current_user.tenantId},
    )
    rows = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"

    _build_header(
        ws,
        [
            "Name",
            "Contact",
            "Email",
            "Phone",
            "Address",
            "Category",
            "Rating",
            "Status",
            "Payment Terms",
        ],
    )

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(wb.active)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vendors_export.xlsx"},
    )


@router.get("/vendors/pdf")
async def export_vendors_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    result = await db.execute(
        text(
            'SELECT name, contact, email, phone, category, rating, status FROM vendors WHERE "tenantId" = :tid ORDER BY name'
        ),
        {"tid": current_user.tenantId},
    )
    rows = result.fetchall()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Vendors Report - Blackbox BOM", styles["Title"]))
    elements.append(Spacer(1, 20))

    data = [["Name", "Contact", "Email", "Phone", "Category", "Rating", "Status"]]
    for row in rows:
        data.append([str(v or "") for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=vendors_report.pdf"},
    )


@router.get("/pos/xlsx")
async def export_pos_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook

    result = await db.execute(
        text(
            'SELECT "poNumber", "poDate", "vendorName", project, "poTotal", status FROM "po_headers" WHERE "tenantId" = :tid ORDER BY "poDate" DESC'
        ),
        {"tid": current_user.tenantId},
    )
    pos = result.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    _build_header(ws, ["PO Number", "Date", "Vendor", "Project", "Total", "Status"])

    for row_idx, row in enumerate(pos, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pos_export.xlsx"},
    )


@router.get("/parts/pdf")
async def export_parts_pdf(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    result = await db.execute(
        text(
            'SELECT pn, name, category, manufacturer, cost, lead, origin, status FROM parts WHERE "tenantId" = :tid ORDER BY pn'
        ),
        {"tid": current_user.tenantId},
    )
    rows = result.fetchall()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Parts Report - Blackbox BOM", styles["Title"]))
    elements.append(Spacer(1, 20))

    data = [["PN", "Name", "Category", "Manufacturer", "Cost", "Lead", "Origin", "Status"]]
    for row in rows:
        data.append([str(v or "") for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=parts_report.pdf"},
    )


@router.get("/summary")
async def export_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tid = current_user.tenantId
    parts_count = await db.execute(
        text('SELECT COUNT(*) FROM parts WHERE "tenantId" = :tid'), {"tid": tid}
    )
    vendors_count = await db.execute(
        text('SELECT COUNT(*) FROM vendors WHERE "tenantId" = :tid'), {"tid": tid}
    )
    po_count = await db.execute(
        text('SELECT COUNT(*) FROM "po_headers" WHERE "tenantId" = :tid'), {"tid": tid}
    )
    docs_count = await db.execute(
        text('SELECT COUNT(*) FROM documents WHERE "tenantId" = :tid'), {"tid": tid}
    )

    return {
        "reportType": "summary",
        "generatedAt": "2026-06-05",
        "totals": {
            "parts": parts_count.scalar() or 0,
            "vendors": vendors_count.scalar() or 0,
            "purchaseOrders": po_count.scalar() or 0,
            "documents": docs_count.scalar() or 0,
        },
    }


@router.get("/bom/xlsx")
async def export_bom_xlsx(
    template_id: int = Query(..., description="BOM Template ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    template = await db.execute(
        text('SELECT * FROM bom_templates WHERE id = :id AND "tenantId" = :tid'),
        {"id": template_id, "tid": current_user.tenantId},
    )
    tmpl = template.mappings().first()
    if not tmpl:
        raise HTTPException(404, "BOM template not found")

    items_result = await db.execute(
        text(
            "SELECT bi.pn, bi.name, bi.quantity, bi.uom, bi.category, bi.vendor, bi.cost "
            'FROM bom_items bi WHERE bi."bomTemplateId" = :tid ORDER BY bi."sortOrder"'
        ),
        {"tid": template_id},
    )
    items = items_result.fetchall()

    total_cost = sum((row[2] or 0) * (row[6] or 0) for row in items)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Items"

    _build_header(
        ws, ["#", "PN", "Name", "Qty", "UoM", "Category", "Vendor", "Unit Cost", "Ext Cost"]
    )

    for i, row in enumerate(items, 1):
        ext = (row[2] or 0) * (row[6] or 0)
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=row[0] or "")
        ws.cell(row=i + 1, column=3, value=row[1] or "")
        ws.cell(row=i + 1, column=4, value=row[2] or 0)
        ws.cell(row=i + 1, column=5, value=row[3] or "")
        ws.cell(row=i + 1, column=6, value=row[4] or "")
        ws.cell(row=i + 1, column=7, value=row[5] or "")
        ws.cell(row=i + 1, column=8, value=row[6] or 0)
        ws.cell(row=i + 1, column=9, value=round(ext, 2))
    total_row = len(items) + 2
    ws.cell(row=total_row, column=7, value="TOTAL")
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=round(total_cost, 2))
    ws.cell(row=total_row, column=9).font = Font(bold=True)

    _auto_width(ws)

    safe_name = (tmpl["name"] or "bom").replace(" ", "_").replace("/", "-")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bom_{safe_name}.xlsx"},
    )


@router.get("/bom/pdf")
async def export_bom_pdf(
    template_id: int = Query(..., description="BOM Template ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    template = await db.execute(
        text('SELECT * FROM bom_templates WHERE id = :id AND "tenantId" = :tid'),
        {"id": template_id, "tid": current_user.tenantId},
    )
    tmpl = template.mappings().first()
    if not tmpl:
        raise HTTPException(404, "BOM template not found")

    items_result = await db.execute(
        text(
            "SELECT bi.pn, bi.name, bi.quantity, bi.uom, bi.category, bi.vendor, bi.cost "
            'FROM bom_items bi WHERE bi."bomTemplateId" = :tid ORDER BY bi."sortOrder"'
        ),
        {"tid": template_id},
    )
    items = items_result.fetchall()

    total_cost = sum((row[2] or 0) * (row[6] or 0) for row in items)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"BOM Report: {tmpl['name']}", styles["Title"]))
    elements.append(Spacer(1, 10))
    elements.append(
        Paragraph(
            f"Status: {tmpl.get('status', 'N/A')} | "
            f"Created: {str(tmpl.get('createdAt', 'N/A'))[:10]} | "
            f"Items: {len(items)}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 20))

    data = [["#", "PN", "Name", "Qty", "UoM", "Category", "Vendor", "Unit Cost", "Ext Cost"]]
    for i, row in enumerate(items, 1):
        ext = (row[2] or 0) * (row[6] or 0)
        data.append(
            [
                str(i),
                str(row[0] or ""),
                str(row[1] or ""),
                str(row[2] or ""),
                str(row[3] or ""),
                str(row[4] or ""),
                str(row[5] or ""),
                f"${(row[6] or 0):,.2f}",
                f"${ext:,.2f}",
            ]
        )
    data.append(["", "", "", "", "", "", "TOTAL", "", f"${total_cost:,.2f}"])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F3F4F6")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E5E7EB")),
                ("FONTSIZE", (0, -1), (-1, -1), 8),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("ALIGN", (7, 0), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    safe_name = (tmpl["name"] or "bom").replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=bom_{safe_name}.pdf"},
    )
